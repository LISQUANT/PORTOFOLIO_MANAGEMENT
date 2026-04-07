import blpapi
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

TICKERS = {
    "Micron":    "MU US Equity",
    "Microsoft": "MSFT US Equity",
    "Eli Lilly": "LLY US Equity",
    "ASML":      "ASML NA Equity",
    "LVMH":      "MC FP Equity",
    "SPY":       "SPX INDEX",    
}

FIELDS = {
    "PX_OPEN":   "Open",
    "PX_HIGH":   "High",
    "PX_LOW":    "Low",
    "PX_LAST":   "Close",
    "PX_VOLUME": "Volume",
}

END_DATE   = datetime.today()
START_DATE = END_DATE - timedelta(days=365 * 10)

OUTPUT_FILE = "OHLCV_10Y.xlsx"

#blpapi session

def start_session():
    opts = blpapi.SessionOptions()
    opts.setServerHost("localhost")
    opts.setServerPort(8194)
    session = blpapi.Session(opts)
    if not session.start():
        print("ERROR: Failed to start Bloomberg session.")
        sys.exit(1)
    if not session.openService("//blp/refdata"):
        print("ERROR: Failed to open //blp/refdata service.")
        sys.exit(1)
    return session


def fetch_historical(session, ticker, start, end):
    service  = session.getService("//blp/refdata")
    request  = service.createRequest("HistoricalDataRequest")

    request.getElement("securities").appendValue(ticker)
    for field in FIELDS.keys():
        request.getElement("fields").appendValue(field)

    request.set("startDate", start.strftime("%Y%m%d"))
    request.set("endDate",   end.strftime("%Y%m%d"))
    request.set("periodicitySelection", "DAILY")
    session.sendRequest(request)

    records = []
    while True:
        event = session.nextEvent(500)
        for msg in event:
            if msg.messageType() == blpapi.Name("HistoricalDataResponse"):
                security_data = msg.getElement("securityData")
                field_data    = security_data.getElement("fieldData")
                for i in range(field_data.numValues()):
                    point = field_data.getValue(i)
                    row   = {"Date": point.getElementAsDatetime("date")}
                    for blp_field, col_name in FIELDS.items():
                        row[col_name] = (
                            point.getElementAsFloat(blp_field)
                            if point.hasElement(blp_field) else None
                        )
                    records.append(row)
        if event.eventType() == blpapi.Event.RESPONSE:
            break

    df = pd.DataFrame(records)
    if not df.empty:
        df["Date"] = pd.to_datetime(df["Date"])
        df = df.sort_values("Date").reset_index(drop=True)
    return df


#Excel formatting 

HEADER_FILL   = PatternFill("solid", start_color="1F3864")   # dark navy
ALT_FILL      = PatternFill("solid", start_color="EBF0FA")   # light blue
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT     = Font(name="Arial", size=10)
CENTER        = Alignment(horizontal="center", vertical="center")
LEFT          = Alignment(horizontal="left",   vertical="center")
THIN          = Side(border_style="thin", color="B8C4D0")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def format_sheet(ws, df, ticker_name):
    ws.freeze_panes = "B2"

    headers = ["Date", "Open", "High", "Low", "Close", "Volume"]
    col_widths = [14, 12, 12, 12, 12, 18]

    # Header row
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()

        date_cell = ws.cell(row=row_idx, column=1, value=row.Date.date())
        date_cell.number_format = "YYYY-MM-DD"
        date_cell.font      = BODY_FONT
        date_cell.fill      = fill
        date_cell.alignment = CENTER
        date_cell.border    = BORDER

        for col_idx, val in enumerate([row.Open, row.High, row.Low, row.Close], start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.number_format = "#,##0.00"
            cell.font      = BODY_FONT
            cell.fill      = fill
            cell.alignment = CENTER
            cell.border    = BORDER

        vol_cell = ws.cell(row=row_idx, column=6, value=row.Volume)
        vol_cell.number_format = "#,##0"
        vol_cell.font      = BODY_FONT
        vol_cell.fill      = fill
        vol_cell.alignment = CENTER
        vol_cell.border    = BORDER

    last_row = len(df) + 3
    note = ws.cell(
        row=last_row,
        column=1,
        value=f"Source: Bloomberg Terminal, {datetime.today().strftime('%d/%m/%Y')}, {TICKERS.get(ticker_name, ticker_name)}"
    )
    note.font = Font(name="Arial", size=8, italic=True, color="808080")


def add_combined_sheet(wb, all_data):
    ws = wb.create_sheet("ALL_COMBINED", 0)

    headers = ["Ticker", "Date", "Open", "High", "Low", "Close", "Volume"]
    col_widths = [14, 14, 12, 12, 12, 12, 18]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "B2"

    row_idx = 2
    for ticker_name, df in all_data.items():
        for row in df.itertuples(index=False):
            fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()

            ws.cell(row=row_idx, column=1, value=ticker_name).font = BODY_FONT
            ws.cell(row=row_idx, column=1).fill      = fill
            ws.cell(row=row_idx, column=1).alignment = CENTER
            ws.cell(row=row_idx, column=1).border    = BORDER

            date_cell = ws.cell(row=row_idx, column=2, value=row.Date.date())
            date_cell.number_format = "YYYY-MM-DD"
            date_cell.font = BODY_FONT; date_cell.fill = fill
            date_cell.alignment = CENTER; date_cell.border = BORDER

            for col_idx, val in enumerate([row.Open, row.High, row.Low, row.Close], start=3):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.number_format = "#,##0.00"
                cell.font = BODY_FONT; cell.fill = fill
                cell.alignment = CENTER; cell.border = BORDER

            vol_cell = ws.cell(row=row_idx, column=7, value=row.Volume)
            vol_cell.number_format = "#,##0"
            vol_cell.font = BODY_FONT; vol_cell.fill = fill
            vol_cell.alignment = CENTER; vol_cell.border = BORDER

            row_idx += 1

def main():
    print("Starting Bloomberg session")
    session = start_session()
    print(f"Fetching data: {START_DATE.strftime('%Y-%m-%d')} → {END_DATE.strftime('%Y-%m-%d')}\n")

    all_data = {}
    writer   = pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl")

    for name, ticker in TICKERS.items():
        print(f"  Fetching {name} ({ticker})")
        df = fetch_historical(session, ticker, START_DATE, END_DATE)
        print(f"    → {len(df)} trading days fetched.")
        df.to_excel(writer, sheet_name=name, index=False)
        all_data[name] = df

    writer.close()
    session.stop()

    print("\nApplying formatting")
    wb = load_workbook(OUTPUT_FILE)

    for name in TICKERS.keys():
        ws = wb[name]
        ws.delete_rows(1, ws.max_row)   
        df = all_data[name]
        format_sheet(ws, df, name)

    add_combined_sheet(wb, all_data)
    wb.save(OUTPUT_FILE)

    print(f"\nDone. File saved: {OUTPUT_FILE}")
    print(f"Sheets: ALL_COMBINED + {', '.join(TICKERS.keys())}")

if __name__ == "__main__":
    main()